import io
from flask import send_file, request, flash, redirect, url_for
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from app.export import bp
from app.auth.routes import manager_required
from app.models import (
    User, UserRole, WorkSchedule, ScheduleShift, AttendanceRecord,
    Payroll, db
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_styled_workbook():
    """Tao workbook voi style mac dinh"""
    wb = Workbook()
    return wb


def style_header_row(ws, row_num, col_count):
    """Style cho header row"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


@bp.route('/schedule-matrix')
@login_required
@manager_required
def export_schedule_matrix():
    """Xuat ma tran lich dang ky"""
    # Lay tuan hien tai hoac tuan duoc chon
    week_str = request.args.get('week')
    if week_str:
        try:
            week_start = datetime.strptime(week_str, '%Y-%m-%d').date()
        except ValueError:
            week_start = date.today() - timedelta(days=date.today().weekday())
    else:
        week_start = date.today() - timedelta(days=date.today().weekday())

    week_end = week_start + timedelta(days=6)

    # Lay danh sach nhan vien
    users = User.query.filter_by(status='active', role=UserRole.STAFF).order_by(User.full_name).all()

    # Tao workbook
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Ma tran lich"

    # Header
    ws['A1'] = f"MA TRAN LICH LAM VIEC: {week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    # Column headers
    headers = ['STT', 'Ho ten', 'Ma NV']
    days = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
    headers.extend(days)

    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    # Data rows
    row = 4
    for idx, user in enumerate(users, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=user.full_name)
        ws.cell(row=row, column=3, value=user.username)

        # Lay lich lam viec cua user trong tuan
        schedule = WorkSchedule.query.filter(
            WorkSchedule.user_id == user.id,
            WorkSchedule.week_start_date == week_start
        ).first()

        if schedule:
            for day_idx in range(7):
                current_date = week_start + timedelta(days=day_idx)
                day_shifts = [s for s in schedule.shifts if s.date == current_date]

                if day_shifts:
                    shift_str = ', '.join([s.shift_type.value[0].upper() for s in day_shifts])
                    ws.cell(row=row, column=4 + day_idx, value=shift_str)

        row += 1

    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ma_tran_lich_{week_start.strftime('%Y%m%d')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/attendance')
@login_required
@manager_required
def export_attendance():
    """Xuat bang cham cong"""
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)

    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1)
    else:
        month_end = date(year, month + 1, 1)

    # Lay danh sach nhan vien
    users = User.query.filter_by(status='active', role=UserRole.STAFF).order_by(User.full_name).all()

    # Tao workbook
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Cham cong"

    # Header
    ws['A1'] = f"BANG CHAM CONG THANG {month}/{year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    # Column headers
    headers = ['STT', 'Ho ten', 'Ma NV', 'Tong ca', 'Tong gio', 'Di muon', 'Ghi chu']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    # Data rows
    row = 4
    for idx, user in enumerate(users, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=user.full_name)
        ws.cell(row=row, column=3, value=user.username)

        # Lay cham cong
        records = AttendanceRecord.query.filter(
            AttendanceRecord.user_id == user.id,
            AttendanceRecord.date >= month_start,
            AttendanceRecord.date < month_end
        ).all()

        total_shifts = len(records)
        total_hours = sum(r.total_work_hours for r in records)
        late_count = sum(1 for r in records if r.is_late)

        ws.cell(row=row, column=4, value=total_shifts)
        ws.cell(row=row, column=5, value=round(total_hours, 1))
        ws.cell(row=row, column=6, value=late_count)

        row += 1

    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"cham_cong_{month}_{year}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/payroll')
@login_required
@manager_required
def export_payroll():
    """Xuat bang luong"""
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)

    # Lay payroll data
    payrolls = db.session.query(Payroll, User).join(User).filter(
        Payroll.month == month,
        Payroll.year == year
    ).order_by(User.full_name).all()

    # Tao workbook
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Bang luong"

    # Header
    ws['A1'] = f"BANG LUONG THANG {month}/{year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:K1')

    # Column headers
    headers = ['STT', 'Ho ten', 'Ma NV', 'So gio', 'Luong gop', 'An ca', 'Thuong',
               'Phat', 'Tam ung', 'Thuc linh', 'Trang thai']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    style_header_row(ws, 3, len(headers))

    # Data rows
    row = 4
    for idx, (payroll, user) in enumerate(payrolls, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=user.full_name)
        ws.cell(row=row, column=3, value=user.username)
        ws.cell(row=row, column=4, value=payroll.total_work_hours)
        ws.cell(row=row, column=5, value=payroll.gross_salary)
        ws.cell(row=row, column=6, value=payroll.meal_support_amount)
        ws.cell(row=row, column=7, value=payroll.total_reward)
        ws.cell(row=row, column=8, value=payroll.total_penalty)
        ws.cell(row=row, column=9, value=payroll.advance_payment)
        ws.cell(row=row, column=10, value=payroll.net_salary)

        status_map = {'draft': 'Chua duyet', 'approved': 'Da duyet', 'paid': 'Da tra'}
        ws.cell(row=row, column=11, value=status_map.get(payroll.status.value, payroll.status.value))

        row += 1

    # Total row
    ws.cell(row=row, column=1, value='TONG')
    ws.cell(row=row, column=5, value=sum(p.gross_salary for p, u in payrolls))
    ws.cell(row=row, column=6, value=sum(p.meal_support_amount for p, u in payrolls))
    ws.cell(row=row, column=7, value=sum(p.total_reward for p, u in payrolls))
    ws.cell(row=row, column=8, value=sum(p.total_penalty for p, u in payrolls))
    ws.cell(row=row, column=9, value=sum(p.advance_payment for p, u in payrolls))
    ws.cell(row=row, column=10, value=sum(p.net_salary for p, u in payrolls))

    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).font = Font(bold=True)

    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"bang_luong_{month}_{year}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
