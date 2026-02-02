"""
Module xu ly import file Excel cham cong tu may van tay

Ho tro 2 format:
1. Format dong (row format):
   | Ma NV | Ho ten | Ngay | Gio vao | Gio ra |

2. Format cot (pivot table format - tu may cham cong):
   | Ma NV | Ten | 01-12 | 02-12 | 03-12 | ...
   Moi cell chua gio vao/ra cach boi xuong dong
"""

import os
from openpyxl import load_workbook
from datetime import datetime, time, timedelta, date as date_type
from app.models import (
    AttendanceRecord, User, WorkSchedule, ScheduleShift,
    ShiftType, db
)


def parse_time(value):
    """Chuyen doi gia tri thanh time object"""
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        # Thu parse cac format khac nhau
        for fmt in ['%H:%M:%S', '%H:%M', '%H.%M']:
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue
    return None


def parse_date(value, date_format='auto', year=None):
    """Chuyen doi gia tri thanh date object

    Args:
        value: Gia tri can parse
        date_format: Dinh dang ngay thang ('auto', 'dd-mm', 'mm-dd', 'dd/mm', 'mm/dd', 'yyyy-mm-dd', 'dd/mm/yyyy')
        year: Nam mac dinh neu khong co trong gia tri
    """
    if year is None:
        year = datetime.now().year

    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date') and callable(getattr(value, 'date', None)):
        return value.date()
    if isinstance(value, date_type):
        return value

    if isinstance(value, str):
        value = value.strip()

        # Dinh dang cu the theo lua chon
        format_map = {
            'dd-mm': [('%d-%m', True)],
            'mm-dd': [('%m-%d', True)],
            'dd/mm': [('%d/%m', True)],
            'mm/dd': [('%m/%d', True)],
            'yyyy-mm-dd': [('%Y-%m-%d', False)],
            'dd/mm/yyyy': [('%d/%m/%Y', False)],
        }

        if date_format != 'auto' and date_format in format_map:
            for fmt, needs_year in format_map[date_format]:
                try:
                    parsed = datetime.strptime(value, fmt)
                    if needs_year:
                        return date_type(year, parsed.month, parsed.day)
                    return parsed.date()
                except ValueError:
                    continue

        # Tu dong nhan dien
        formats = [
            ('%Y-%m-%d', False),
            ('%d/%m/%Y', False),
            ('%d-%m-%Y', False),
            ('%m/%d/%Y', False),
            ('%d-%m', True),
            ('%d/%m', True),
        ]
        for fmt, needs_year in formats:
            try:
                parsed = datetime.strptime(value, fmt)
                if needs_year:
                    return date_type(year, parsed.month, parsed.day)
                return parsed.date()
            except ValueError:
                continue
    return None


def calculate_late_minutes(scheduled_start, actual_checkin):
    """
    Tinh so phut di muon
    Returns: So phut di muon (0 neu dung gio hoac som)
    """
    if actual_checkin is None or scheduled_start is None:
        return 0

    if actual_checkin <= scheduled_start:
        return 0

    # Chuyen sang datetime de tinh
    today = datetime.today().date()
    scheduled_dt = datetime.combine(today, scheduled_start)
    actual_dt = datetime.combine(today, actual_checkin)

    delta = actual_dt - scheduled_dt
    return int(delta.total_seconds() / 60)


def round_up_to_half_hour(t):
    """
    Lam tron len 30 phut
    VD: 7h06 -> 7h30, 7h35 -> 8h00
    """
    if t is None:
        return None

    minutes = t.minute
    if minutes <= 30:
        new_minute = 30
        new_hour = t.hour
    else:
        new_minute = 0
        new_hour = t.hour + 1
        if new_hour >= 24:
            new_hour = 23
            new_minute = 59

    return time(new_hour, new_minute, 0)


def calculate_work_hours(scheduled_start, scheduled_end, actual_checkin, late_minutes):
    """
    Tinh gio lam viec thuc te
    - Neu dung gio: Tinh theo lich
    - Neu di muon: Lam tron xuong
    """
    if scheduled_start is None or scheduled_end is None:
        return 0

    today = datetime.today().date()
    start_dt = datetime.combine(today, scheduled_start)
    end_dt = datetime.combine(today, scheduled_end)

    if late_minutes == 0:
        # Tinh theo lich
        delta = end_dt - start_dt
    else:
        # Lam tron len thoi gian bat dau
        rounded_start = round_up_to_half_hour(actual_checkin)
        if rounded_start:
            start_dt = datetime.combine(today, rounded_start)
        delta = end_dt - start_dt

    hours = delta.total_seconds() / 3600
    return max(0, hours)


def find_scheduled_shift(user_id, date, actual_checkin):
    """Tim ca lam viec da duoc phan cho NV"""
    # Lay tat ca shifts cua user trong ngay do
    shifts = db.session.query(ScheduleShift).join(WorkSchedule).filter(
        WorkSchedule.user_id == user_id,
        ScheduleShift.date == date,
        ScheduleShift.is_confirmed == True
    ).all()

    if not shifts:
        return None

    if len(shifts) == 1:
        return shifts[0]

    # Neu co nhieu ca, tim ca gan nhat voi thoi gian check-in
    if actual_checkin:
        closest_shift = None
        min_diff = float('inf')

        for shift in shifts:
            diff = abs(
                (datetime.combine(date, actual_checkin) -
                 datetime.combine(date, shift.shift_start_time)).total_seconds()
            )
            if diff < min_diff:
                min_diff = diff
                closest_shift = shift

        return closest_shift

    return shifts[0]


def load_excel_file(file_path):
    """
    Load file Excel (.xlsx hoac .xls)
    Returns: (workbook_data, is_openpyxl, error_message)
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.xlsx':
        try:
            wb = load_workbook(file_path)
            return wb.active, True, None
        except Exception as e:
            return None, None, f'Khong the doc file xlsx: {str(e)}'

    elif ext == '.xls':
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            return ws, False, None
        except ImportError:
            return None, None, 'Can cai dat xlrd de doc file .xls'
        except Exception as e:
            return None, None, f'Khong the doc file xls: {str(e)}'

    else:
        return None, None, f'Dinh dang file khong ho tro: {ext}'


def get_cell_value(ws, row, col, is_openpyxl):
    """Lay gia tri cell tu worksheet"""
    if is_openpyxl:
        return ws.cell(row=row+1, column=col+1).value  # openpyxl 1-indexed
    else:
        try:
            return ws.cell_value(row, col)  # xlrd 0-indexed
        except:
            return None


def get_sheet_dimensions(ws, is_openpyxl):
    """Lay kich thuoc sheet (rows, cols)"""
    if is_openpyxl:
        return ws.max_row, ws.max_column
    else:
        return ws.nrows, ws.ncols


def detect_format(ws, is_openpyxl):
    """
    Phat hien format file Excel
    Returns: 'row' hoac 'pivot'
    """
    # Lay header row
    max_cols = get_sheet_dimensions(ws, is_openpyxl)[1]

    # Kiem tra cot thu 3 (index 2)
    if max_cols >= 3:
        header_val = get_cell_value(ws, 0, 2, is_openpyxl)
        if header_val:
            header_str = str(header_val).strip().lower()
            # Neu header la 'ngay' -> row format
            if header_str in ['ngay', 'ngày', 'date']:
                return 'row'
            # Neu header la ngay thang (12-01, 01/12, etc) -> pivot format
            if '-' in str(header_val) or '/' in str(header_val):
                return 'pivot'

    return 'row'  # Default


def parse_pivot_date(header_val, year=None):
    """
    Parse date tu header cot pivot format
    VD: '01-12', '12-01', '01/12', '12/01'
    """
    if year is None:
        year = datetime.now().year

    if header_val is None:
        return None

    header_str = str(header_val).strip()

    # Thu cac format khac nhau
    for sep in ['-', '/']:
        if sep in header_str:
            parts = header_str.split(sep)
            if len(parts) == 2:
                try:
                    # Thu DD-MM truoc
                    day = int(parts[0])
                    month = int(parts[1])
                    if 1 <= day <= 31 and 1 <= month <= 12:
                        return date_type(year, month, day)
                except ValueError:
                    pass
                try:
                    # Thu MM-DD
                    month = int(parts[0])
                    day = int(parts[1])
                    if 1 <= day <= 31 and 1 <= month <= 12:
                        return date_type(year, month, day)
                except ValueError:
                    pass

    return None


def parse_checkin_checkout(cell_value):
    """
    Parse gio check-in va check-out tu cell
    Cell co the chua:
    - 1 thoi gian: '7:00'
    - 2 thoi gian cach newline: '7:00\n12:00'
    - 2 thoi gian cach space: '7:00 12:00'

    Returns: (checkin_time, checkout_time)
    """
    if cell_value is None:
        return None, None

    cell_str = str(cell_value).strip()
    if not cell_str:
        return None, None

    # Tach cac phan tu
    times = []
    for sep in ['\n', '\r\n', '  ', ' - ', '/']:
        if sep in cell_str:
            parts = cell_str.split(sep)
            for p in parts:
                p = p.strip()
                if p:
                    t = parse_time(p)
                    if t:
                        times.append(t)
            break

    if not times:
        # Chi co 1 gia tri
        t = parse_time(cell_str)
        if t:
            times = [t]

    if len(times) == 0:
        return None, None
    elif len(times) == 1:
        return times[0], None
    else:
        return times[0], times[1]


def import_pivot_format(ws, is_openpyxl, date_format='auto'):
    """Import file Excel format pivot (ngay o cot)"""
    records_created = 0
    errors = []
    records = []

    max_rows, max_cols = get_sheet_dimensions(ws, is_openpyxl)
    current_year = datetime.now().year

    # Parse headers (dates)
    date_columns = {}  # {col_index: date}
    for col in range(2, max_cols):
        header = get_cell_value(ws, 0, col, is_openpyxl)
        if header:
            parsed_date = parse_pivot_date(header, current_year)
            if parsed_date:
                date_columns[col] = parsed_date

    if not date_columns:
        return {'success': 0, 'errors': ['Khong tim thay cot ngay thang trong file'], 'records': []}

    # Process each employee row
    for row in range(1, max_rows):
        employee_code = get_cell_value(ws, row, 0, is_openpyxl)
        if not employee_code:
            continue

        employee_code = str(employee_code).strip()

        # Tim user
        user = User.query.filter_by(username=employee_code).first()
        if not user:
            user = User.query.filter(User.full_name.ilike(f'%{employee_code}%')).first()

        if not user:
            # Bo qua NV khong tim thay (khong bao loi)
            continue

        # Process each date column
        for col, date in date_columns.items():
            cell_value = get_cell_value(ws, row, col, is_openpyxl)
            if not cell_value:
                continue

            checkin_time, checkout_time = parse_checkin_checkout(cell_value)
            if not checkin_time:
                continue

            try:
                # Tim ca lam viec da duoc phan
                scheduled_shift = find_scheduled_shift(user.id, date, checkin_time)

                # Neu khong co lich, tao record voi shift type suy doan
                if not scheduled_shift:
                    # Suy doan shift type tu thoi gian check-in
                    if checkin_time.hour < 12:
                        shift_type = ShiftType.MORNING
                        scheduled_start = time(7, 0)
                        scheduled_end = time(12, 0)
                    elif checkin_time.hour < 18:
                        shift_type = ShiftType.AFTERNOON
                        scheduled_start = time(12, 0)
                        scheduled_end = time(18, 0)
                    else:
                        shift_type = ShiftType.EVENING
                        scheduled_start = time(18, 0)
                        scheduled_end = time(22, 0)
                else:
                    shift_type = scheduled_shift.shift_type
                    scheduled_start = scheduled_shift.shift_start_time
                    scheduled_end = scheduled_shift.shift_end_time

                # Kiem tra da import chua
                existing = AttendanceRecord.query.filter_by(
                    user_id=user.id,
                    date=date,
                    shift_type=shift_type
                ).first()

                if existing:
                    continue  # Skip quietly for pivot format

                # Tinh toan
                late_minutes = calculate_late_minutes(scheduled_start, checkin_time)
                work_hours = calculate_work_hours(
                    scheduled_start, scheduled_end, checkin_time, late_minutes
                )

                # Check early bird
                early_bird_threshold = time(6, 55)
                is_early_bird = checkin_time and checkin_time < early_bird_threshold

                # Tao record
                record = AttendanceRecord(
                    user_id=user.id,
                    date=date,
                    shift_type=shift_type,
                    scheduled_start=scheduled_start,
                    scheduled_end=scheduled_end,
                    actual_checkin=checkin_time,
                    actual_checkout=checkout_time,
                    late_minutes=late_minutes,
                    total_work_hours=round(work_hours, 2),
                    is_late=(late_minutes > 0),
                    is_early_bird=is_early_bird
                )

                db.session.add(record)
                records.append(record)
                records_created += 1

            except Exception as e:
                errors.append(f'NV {employee_code} ngay {date}: Loi - {str(e)}')

    if records_created > 0:
        db.session.commit()

    return {
        'success': records_created,
        'errors': errors,
        'records': records
    }


def import_attendance_excel(file_path, date_format='auto'):
    """
    Import file Excel cham cong

    Args:
        file_path: Duong dan file Excel
        date_format: Dinh dang ngay thang ('auto', 'dd-mm', 'mm-dd', etc.)

    Returns:
        dict: {
            'success': so record thanh cong,
            'errors': list loi,
            'records': list records da tao
        }
    """
    ws, is_openpyxl, error = load_excel_file(file_path)
    if error:
        return {'success': 0, 'errors': [error], 'records': []}

    # Detect format
    file_format = detect_format(ws, is_openpyxl)

    if file_format == 'pivot':
        return import_pivot_format(ws, is_openpyxl, date_format)

    # Row format (original)
    if is_openpyxl:
        return import_row_format_openpyxl(ws, date_format)
    else:
        return import_row_format_xlrd(ws, date_format)


def import_row_format_xlrd(ws, date_format='auto'):
    """Import file xls row format using xlrd"""
    records_created = 0
    errors = []
    records = []

    for row_num in range(1, ws.nrows):
        try:
            row = [ws.cell_value(row_num, col) for col in range(ws.ncols)]

            if not row or not row[0]:
                continue

            employee_code = str(row[0]).strip()
            date = parse_date(row[2], date_format) if len(row) > 2 else None
            checkin_time = parse_time(row[3]) if len(row) > 3 else None
            checkout_time = parse_time(row[4]) if len(row) > 4 else None

            if not date:
                errors.append(f'Dong {row_num+1}: Ngay khong hop le')
                continue

            user = User.query.filter_by(username=employee_code).first()
            if not user:
                user = User.query.filter(User.full_name.ilike(f'%{employee_code}%')).first()

            if not user:
                # Bo qua NV khong tim thay
                continue

            scheduled_shift = find_scheduled_shift(user.id, date, checkin_time)

            if not scheduled_shift:
                # Bo qua neu khong co lich
                continue

            existing = AttendanceRecord.query.filter_by(
                user_id=user.id,
                date=date,
                shift_type=scheduled_shift.shift_type
            ).first()

            if existing:
                # Bo qua neu da import
                continue

            late_minutes = calculate_late_minutes(
                scheduled_shift.shift_start_time,
                checkin_time
            )

            work_hours = calculate_work_hours(
                scheduled_shift.shift_start_time,
                scheduled_shift.shift_end_time,
                checkin_time,
                late_minutes
            )

            early_bird_threshold = time(6, 55)
            is_early_bird = checkin_time and checkin_time < early_bird_threshold

            record = AttendanceRecord(
                user_id=user.id,
                date=date,
                shift_type=scheduled_shift.shift_type,
                scheduled_start=scheduled_shift.shift_start_time,
                scheduled_end=scheduled_shift.shift_end_time,
                actual_checkin=checkin_time,
                actual_checkout=checkout_time,
                late_minutes=late_minutes,
                total_work_hours=round(work_hours, 2),
                is_late=(late_minutes > 0),
                is_early_bird=is_early_bird
            )

            db.session.add(record)
            records.append(record)
            records_created += 1

        except Exception as e:
            errors.append(f'Dong {row_num+1}: Loi - {str(e)}')

    if records_created > 0:
        db.session.commit()

    return {
        'success': records_created,
        'errors': errors,
        'records': records
    }


def import_row_format_openpyxl(ws, date_format='auto'):
    """Import file xlsx row format using openpyxl (original logic)"""

    records_created = 0
    errors = []
    records = []

    # Bo qua dong header
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Kiem tra dong trong
            if not row or not row[0]:
                continue

            employee_code = str(row[0]).strip()
            # employee_name = row[1]  # Khong can dung
            date = parse_date(row[2], date_format)
            checkin_time = parse_time(row[3])
            checkout_time = parse_time(row[4]) if len(row) > 4 else None

            if not date:
                errors.append(f'Dong {row_num}: Ngay khong hop le')
                continue

            # Tim user theo username
            user = User.query.filter_by(username=employee_code).first()
            if not user:
                # Thu tim theo full_name
                user = User.query.filter(User.full_name.ilike(f'%{employee_code}%')).first()

            if not user:
                # Bo qua NV khong tim thay
                continue

            # Tim ca lam viec da duoc phan
            scheduled_shift = find_scheduled_shift(user.id, date, checkin_time)

            if not scheduled_shift:
                # Bo qua neu khong co lich
                continue

            # Kiem tra da import chua
            existing = AttendanceRecord.query.filter_by(
                user_id=user.id,
                date=date,
                shift_type=scheduled_shift.shift_type
            ).first()

            if existing:
                # Bo qua neu da import
                continue

            # Tinh toan
            late_minutes = calculate_late_minutes(
                scheduled_shift.shift_start_time,
                checkin_time
            )

            work_hours = calculate_work_hours(
                scheduled_shift.shift_start_time,
                scheduled_shift.shift_end_time,
                checkin_time,
                late_minutes
            )

            # Check early bird (di truoc 6h55)
            early_bird_threshold = time(6, 55)
            is_early_bird = checkin_time and checkin_time < early_bird_threshold

            # Tao record
            record = AttendanceRecord(
                user_id=user.id,
                date=date,
                shift_type=scheduled_shift.shift_type,
                scheduled_start=scheduled_shift.shift_start_time,
                scheduled_end=scheduled_shift.shift_end_time,
                actual_checkin=checkin_time,
                actual_checkout=checkout_time,
                late_minutes=late_minutes,
                total_work_hours=round(work_hours, 2),
                is_late=(late_minutes > 0),
                is_early_bird=is_early_bird
            )

            db.session.add(record)
            records.append(record)
            records_created += 1

        except Exception as e:
            errors.append(f'Dong {row_num}: Loi - {str(e)}')

    if records_created > 0:
        db.session.commit()

    return {
        'success': records_created,
        'errors': errors,
        'records': records
    }


# =============================================================================
# PREVIEW IMPORT FUNCTIONS
# =============================================================================

def parse_attendance_preview(file_path, year, date_format='auto'):
    """
    Parse file Excel de preview - khong luu vao DB

    Args:
        file_path: Duong dan file Excel
        year: Nam de ghi date (file co the chi co dd/mm)
        date_format: Dinh dang ngay

    Returns:
        List[dict]: Danh sach records de preview
    """
    ws, is_openpyxl, error = load_excel_file(file_path)
    if error:
        return []

    records = []
    max_rows, max_cols = get_sheet_dimensions(ws, is_openpyxl)

    # Detect format
    file_format = detect_format(ws, is_openpyxl)

    if file_format == 'pivot':
        # Parse pivot format
        records = parse_pivot_preview(ws, is_openpyxl, year, date_format)
    else:
        # Parse row format
        records = parse_row_preview(ws, is_openpyxl, year, date_format)

    return records


def parse_row_preview(ws, is_openpyxl, year, date_format='auto'):
    """Parse file Excel row format de preview"""
    records = []
    max_rows, max_cols = get_sheet_dimensions(ws, is_openpyxl)

    # Tim dong header
    header_row = 0
    for row in range(min(5, max_rows)):
        val = get_cell_value(ws, row, 0, is_openpyxl)
        if val:
            val_str = str(val).lower().strip()
            if val_str in ['ma nv', 'ma', 'code', 'manv', 'stt', 'mã nv', 'mã']:
                header_row = row
                break

    # Parse data rows
    for row in range(header_row + 1, max_rows):
        try:
            # Lay gia tri cac cot
            col0 = get_cell_value(ws, row, 0, is_openpyxl)
            col1 = get_cell_value(ws, row, 1, is_openpyxl)
            col2 = get_cell_value(ws, row, 2, is_openpyxl)
            col3 = get_cell_value(ws, row, 3, is_openpyxl)
            col4 = get_cell_value(ws, row, 4, is_openpyxl)

            # Neu tat ca empty -> skip
            if not any([col0, col1, col2, col3]):
                continue

            # Xac dinh cot nao la gi
            # Truong hop 1: Ma NV, Ho ten, Ngay, Gio vao, Gio ra
            # Truong hop 2: STT, Ho ten, Ngay, Gio vao, Gio ra (khong co ma NV)
            employee_code = None
            full_name = None
            date_val = None
            checkin = None
            checkout = None

            # Thu tim ho ten (cot co chu cai)
            for i, val in enumerate([col0, col1, col2, col3, col4]):
                if val and isinstance(val, str) and len(val) > 2 and not parse_date(val, date_format, year) and not parse_time(val):
                    if i == 0:
                        # Cot 0 la STT hoac Ma NV
                        try:
                            int(str(col0))
                            # La so -> STT, ho ten o cot 1
                            full_name = str(col1).strip() if col1 else None
                            date_val = col2
                            checkin = col3
                            checkout = col4
                        except:
                            # La ma NV
                            employee_code = str(col0).strip()
                            full_name = str(col1).strip() if col1 else None
                            date_val = col2
                            checkin = col3
                            checkout = col4
                    elif i == 1:
                        full_name = str(val).strip()
                        # Cot 0 la ma NV hoac STT
                        try:
                            int(str(col0))
                            # La STT, khong co ma NV
                            pass
                        except:
                            employee_code = str(col0).strip()
                        date_val = col2
                        checkin = col3
                        checkout = col4
                    break

            if not full_name:
                continue

            # Parse date
            parsed_date = parse_date(date_val, date_format, year)
            if not parsed_date:
                continue

            # Parse time
            checkin_time = parse_time(checkin)
            checkout_time = parse_time(checkout)

            # Thu tim user tu employee_code hoac full_name
            matched_user = None
            if employee_code:
                matched_user = User.query.filter_by(username=employee_code).first()

            if not matched_user and full_name:
                matched_user = User.query.filter(User.full_name.ilike(f'%{full_name}%')).first()

            records.append({
                'employee_code': employee_code or (matched_user.username if matched_user else ''),
                'full_name': full_name,
                'date': parsed_date.strftime('%d/%m/%Y'),
                'checkin': checkin_time.strftime('%H:%M') if checkin_time else '',
                'checkout': checkout_time.strftime('%H:%M') if checkout_time else '',
                'matched_user_id': matched_user.id if matched_user else None,
                'matched_user_name': matched_user.full_name if matched_user else None,
                'error': None if matched_user else 'Khong tim thay NV'
            })

        except Exception as e:
            continue

    return records


def parse_pivot_preview(ws, is_openpyxl, year, date_format='auto'):
    """Parse file Excel pivot format de preview"""
    records = []
    max_rows, max_cols = get_sheet_dimensions(ws, is_openpyxl)

    # Parse headers (dates)
    date_columns = {}
    for col in range(2, max_cols):
        header = get_cell_value(ws, 0, col, is_openpyxl)
        if header:
            parsed_date = parse_pivot_date(header, year)
            if parsed_date:
                date_columns[col] = parsed_date

    if not date_columns:
        return []

    # Parse each row
    for row in range(1, max_rows):
        employee_code = get_cell_value(ws, row, 0, is_openpyxl)
        full_name = get_cell_value(ws, row, 1, is_openpyxl)

        if not employee_code and not full_name:
            continue

        employee_code = str(employee_code).strip() if employee_code else ''
        full_name = str(full_name).strip() if full_name else ''

        # Tim user
        matched_user = None
        if employee_code:
            matched_user = User.query.filter_by(username=employee_code).first()
        if not matched_user and full_name:
            matched_user = User.query.filter(User.full_name.ilike(f'%{full_name}%')).first()

        # Parse each date column
        for col, date in date_columns.items():
            cell_value = get_cell_value(ws, row, col, is_openpyxl)
            if not cell_value:
                continue

            checkin_time, checkout_time = parse_checkin_checkout(cell_value)
            if not checkin_time:
                continue

            records.append({
                'employee_code': employee_code or (matched_user.username if matched_user else ''),
                'full_name': full_name or (matched_user.full_name if matched_user else ''),
                'date': date.strftime('%d/%m/%Y'),
                'checkin': checkin_time.strftime('%H:%M') if checkin_time else '',
                'checkout': checkout_time.strftime('%H:%M') if checkout_time else '',
                'matched_user_id': matched_user.id if matched_user else None,
                'matched_user_name': matched_user.full_name if matched_user else None,
                'error': None if matched_user else 'Khong tim thay NV'
            })

    return records


def save_attendance_from_preview(records):
    """
    Luu records tu preview vao database

    Args:
        records: List[dict] tu form preview

    Returns:
        dict: {'success': int, 'errors': list}
    """
    success_count = 0
    errors = []

    for idx, record in enumerate(records):
        try:
            employee_code = record.get('employee_code', '').strip()
            full_name = record.get('full_name', '').strip()
            date_str = record.get('date', '')
            checkin_str = record.get('checkin', '')
            checkout_str = record.get('checkout', '')

            if not date_str:
                errors.append(f'Dong {idx+1}: Thieu ngay')
                continue

            # Tim user
            user = None
            if employee_code:
                user = User.query.filter_by(username=employee_code).first()

            if not user and full_name:
                user = User.query.filter(User.full_name.ilike(f'%{full_name}%')).first()

            if not user:
                errors.append(f'Dong {idx+1}: Khong tim thay NV "{employee_code or full_name}"')
                continue

            # Parse date
            try:
                parsed_date = datetime.strptime(date_str, '%d/%m/%Y').date()
            except:
                errors.append(f'Dong {idx+1}: Ngay khong hop le "{date_str}"')
                continue

            # Parse time
            checkin_time = None
            checkout_time = None
            if checkin_str:
                checkin_time = parse_time(checkin_str)
            if checkout_str:
                checkout_time = parse_time(checkout_str)

            # Tim ca lam viec
            scheduled_shift = find_scheduled_shift(user.id, parsed_date, checkin_time)

            # Neu khong co lich, suy doan shift type
            if not scheduled_shift:
                if checkin_time:
                    if checkin_time.hour < 12:
                        shift_type = ShiftType.MORNING
                        scheduled_start = time(7, 0)
                        scheduled_end = time(12, 0)
                    elif checkin_time.hour < 18:
                        shift_type = ShiftType.AFTERNOON
                        scheduled_start = time(12, 0)
                        scheduled_end = time(18, 0)
                    else:
                        shift_type = ShiftType.EVENING
                        scheduled_start = time(18, 0)
                        scheduled_end = time(22, 0)
                else:
                    shift_type = ShiftType.MORNING
                    scheduled_start = time(7, 0)
                    scheduled_end = time(12, 0)
            else:
                shift_type = scheduled_shift.shift_type
                scheduled_start = scheduled_shift.shift_start_time
                scheduled_end = scheduled_shift.shift_end_time

            # Kiem tra da ton tai chua
            existing = AttendanceRecord.query.filter_by(
                user_id=user.id,
                date=parsed_date,
                shift_type=shift_type
            ).first()

            if existing:
                # Cap nhat ban ghi cu
                existing.actual_checkin = checkin_time
                existing.actual_checkout = checkout_time
                existing.late_minutes = calculate_late_minutes(scheduled_start, checkin_time)
                existing.total_work_hours = round(calculate_work_hours(
                    scheduled_start, scheduled_end, checkin_time, existing.late_minutes
                ), 2)
                existing.is_late = existing.late_minutes > 0
                existing.is_early_bird = checkin_time and checkin_time < time(6, 55)
            else:
                # Tao ban ghi moi
                late_minutes = calculate_late_minutes(scheduled_start, checkin_time)
                work_hours = calculate_work_hours(scheduled_start, scheduled_end, checkin_time, late_minutes)

                attendance = AttendanceRecord(
                    user_id=user.id,
                    date=parsed_date,
                    shift_type=shift_type,
                    scheduled_start=scheduled_start,
                    scheduled_end=scheduled_end,
                    actual_checkin=checkin_time,
                    actual_checkout=checkout_time,
                    late_minutes=late_minutes,
                    total_work_hours=round(work_hours, 2),
                    is_late=late_minutes > 0,
                    is_early_bird=checkin_time and checkin_time < time(6, 55)
                )
                db.session.add(attendance)

            success_count += 1

        except Exception as e:
            errors.append(f'Dong {idx+1}: Loi - {str(e)}')

    if success_count > 0:
        db.session.commit()

    return {
        'success': success_count,
        'errors': errors
    }
